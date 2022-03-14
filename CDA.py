"""
Projeto de Controle de Datas Automático

João Vitor Soares Carneiro
"""
import asyncio
import datetime
import time
from openpyxl import load_workbook
import CDATelBot

# Checagem da Data ##
ano = time.localtime().tm_year
mes = time.localtime().tm_mon
dia = time.localtime().tm_mday
d_recol = 2

data = datetime.datetime(year=ano, month=mes, day=dia + d_recol)
######################

# Criação da planilha
wb = load_workbook(filename='Data_Validades.xlsx')
ws = wb.active

# Leitura da planilha


def __checardata__():
    global cell, produto, count, CVcount
    count = 0
    for row in ws.iter_rows(min_col=2, max_col=2, max_row=15, values_only=True):
        for cell in row:
            count += 1
            pass
        if data == cell:
            CVcount = str(count)
            produto = ws["A"+CVcount].value
            asyncio.run(CDATelBot.start(text=f"{produto} vence em: {cell}"))


# Depuração ##########

# print(data)
# wb.save("debug.xlsx")
if __name__ == '__main__':
    __checardata__()